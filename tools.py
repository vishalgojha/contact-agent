"""Tool functions used by the Contact Name Recovery Agent."""
# contact_name_agent/tools.py

from __future__ import annotations

import csv
import json
import re
import shutil
import subprocess
from pathlib import Path
from typing import Any, Callable

from pydantic import BaseModel, Field

from memory import ContactRecord, MemoryStore
from utils import (
    PHONE_COLUMN_HINTS,
    clamp_confidence,
    dedupe_strings,
    extract_phone_candidates,
    normalize_phone,
    utc_now_iso,
)

try:
    from duckduckgo_search import DDGS
except ImportError:  # pragma: no cover - optional dependency at runtime
    DDGS = None

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency at runtime
    load_workbook = None


NAME_PATTERN = re.compile(r"\b[A-Z][a-z]+(?:\s+[A-Z][a-z]+){0,2}\b")
GENERIC_NAME_TERMS = {
    "Contact",
    "Details",
    "Phone",
    "Number",
    "Mobile",
    "WhatsApp",
    "Broker",
    "Agent",
    "India",
    "Mumbai",
    "Pune",
    "Property",
    "Realty",
    "Listing",
}


class LoadedContact(BaseModel):
    """Normalized phone input record."""

    phone: str
    original_value: str
    source_file: str
    row_number: int | None = None


class WebSnippet(BaseModel):
    """Small public search result fragment."""

    title: str = ""
    snippet: str = ""
    url: str | None = None


class LoadContactsInput(BaseModel):
    """Schema for the load_contacts tool."""

    input_path: str
    default_region: str = "IN"


class MatchGoogleContactsInput(BaseModel):
    """Schema for the match_google_contacts tool."""

    file_path: str
    default_region: str = "IN"


class SearchWebInput(BaseModel):
    """Schema for the search_web tool."""

    phone: str
    max_results: int = Field(default=5, ge=1, le=10)


class ExtractNamesInput(BaseModel):
    """Schema for the extract_names_from_snippets tool."""

    phone: str
    snippets: list[str]


class AskUserInput(BaseModel):
    """Schema for the ask_user tool."""

    question: str


class FinalizeInput(BaseModel):
    """Schema for the finalize tool."""

    phone: str
    name: str
    confidence: float = Field(ge=0.0, le=1.0)
    source: str
    reasoning: str


class TruecallerSearchInput(BaseModel):
    """Schema for the search_truecaller tool."""

    phone: str


class TruecallerSearchResult(BaseModel):
    """Structured result from the vendored truecaller-cli wrapper."""

    name: str
    email: str | None = None
    raw_output: str = ""


class TruecallerStatus(BaseModel):
    """Status for the local truecaller-cli wrapper."""

    ready: bool
    reason: str = ""


class ToolDefinition(BaseModel):
    """Compact tool description shown to the planner."""

    name: str
    description: str
    input_schema: dict[str, Any]


def build_tool_catalog() -> list[dict[str, Any]]:
    """Return the planner-facing tool metadata."""
    tools = [
        ToolDefinition(
            name="load_contacts",
            description="Load numbers from a CSV, TXT, JSON, VCF, or directory and normalize them to E.164.",
            input_schema=LoadContactsInput.model_json_schema(),
        ),
        ToolDefinition(
            name="match_google_contacts",
            description="Look for direct local matches inside a Google Contacts CSV or vCard export.",
            input_schema=MatchGoogleContactsInput.model_json_schema(),
        ),
        ToolDefinition(
            name="search_web",
            description="Perform an opt-in quoted-number public search and return snippets. Use sparingly.",
            input_schema=SearchWebInput.model_json_schema(),
        ),
        ToolDefinition(
            name="extract_names_from_snippets",
            description="Convert noisy snippets into likely contact names using heuristics and optionally the local LLM.",
            input_schema=ExtractNamesInput.model_json_schema(),
        ),
        ToolDefinition(
            name="ask_user",
            description="Ask the user a short console question when confidence is low or evidence conflicts.",
            input_schema=AskUserInput.model_json_schema(),
        ),
        ToolDefinition(
            name="finalize",
            description="Commit the best available name, confidence, source, and reasoning into memory.",
            input_schema=FinalizeInput.model_json_schema(),
        ),
        ToolDefinition(
            name="search_truecaller",
            description="Query the vendored truecaller-cli for an Indian mobile number after OTP setup.",
            input_schema=TruecallerSearchInput.model_json_schema(),
        ),
    ]
    return [tool.model_dump() for tool in tools]


def load_contacts(input_path: str, default_region: str = "IN") -> list[LoadedContact]:
    """Load and normalize contact numbers from one file or a directory."""
    path = Path(input_path)
    if not path.exists():
        raise FileNotFoundError(f"Input path does not exist: {path}")

    files = [path] if path.is_file() else sorted(candidate for candidate in path.iterdir() if candidate.is_file())
    contacts: list[LoadedContact] = []
    seen: set[str] = set()

    for file_path in files:
        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            discovered = _load_contacts_from_csv(file_path, default_region)
        elif suffix == ".xlsx":
            discovered = _load_contacts_from_xlsx(file_path, default_region)
        elif suffix == ".txt":
            discovered = _load_contacts_from_txt(file_path, default_region)
        elif suffix == ".json":
            discovered = _load_contacts_from_json(file_path, default_region)
        elif suffix == ".vcf":
            discovered = _load_contacts_from_vcf(file_path, default_region)
        else:
            continue

        for item in discovered:
            if item.phone in seen:
                continue
            seen.add(item.phone)
            contacts.append(item)

    return contacts


def match_google_contacts(file_path: str, default_region: str = "IN") -> dict[str, str]:
    """Load a local Google Contacts CSV or vCard and return phone-to-name matches."""
    path = Path(file_path)
    if not path.exists():
        return {}

    if path.suffix.lower() == ".xlsx":
        return _load_google_contacts_from_xlsx(path, default_region)
    if path.suffix.lower() == ".csv":
        return _load_google_contacts_from_csv(path, default_region)
    if path.suffix.lower() == ".vcf":
        return _load_google_contacts_from_vcf(path, default_region)
    return {}


def search_web(phone: str, max_results: int = 5) -> list[WebSnippet]:
    """Perform a quoted-number web search and return compact snippets."""
    if DDGS is None:
        raise RuntimeError("duckduckgo-search is not installed. Install requirements.txt first.")

    queries = [f"\"{phone}\"", f"\"{phone.lstrip('+')}\""]
    snippets: list[WebSnippet] = []
    seen: set[str] = set()

    with DDGS() as ddgs:
        for query in queries:
            for hit in ddgs.text(query, max_results=max_results):
                title = str(hit.get("title") or "")
                body = str(hit.get("body") or hit.get("snippet") or "")
                url = str(hit.get("href") or hit.get("url") or "") or None
                dedupe_key = url or f"{title}|{body}"
                if dedupe_key in seen:
                    continue
                seen.add(dedupe_key)
                snippets.append(WebSnippet(title=title, snippet=body, url=url))
                if len(snippets) >= max_results:
                    return snippets

    return snippets


def get_truecaller_status(repo_path: str | Path) -> TruecallerStatus:
    """Check whether the vendored truecaller-cli is ready for lookups."""
    repo = Path(repo_path)
    if shutil.which("node") is None:
        return TruecallerStatus(ready=False, reason="Node.js is not installed or not on PATH.")
    if not repo.exists():
        return TruecallerStatus(ready=False, reason=f"truecaller-cli repo not found at {repo}.")
    if not (repo / "bin" / "app.js").exists():
        return TruecallerStatus(ready=False, reason="truecaller-cli entrypoint is missing.")
    if not (repo / "node_modules").exists():
        return TruecallerStatus(ready=False, reason="truecaller-cli dependencies are not installed. Run npm install in _vendor/truecaller-cli.")
    if not (repo / "config.json").exists():
        return TruecallerStatus(
            ready=False,
            reason=(
                "truecaller-cli is not configured yet. Run `python agent.py setup-truecaller --number YOUR_10_DIGIT_NUMBER` "
                "or `python agent.py setup-truecaller-token --installation-id YOUR_TOKEN` first."
            ),
        )
    return TruecallerStatus(ready=True)


def run_truecaller_register(number: str, repo_path: str | Path) -> int:
    """Run the vendored truecaller-cli registration flow interactively."""
    repo = Path(repo_path)
    mobile = _format_truecaller_number(number)
    if mobile is None:
        raise ValueError("truecaller-cli only supports 10-digit Indian mobile numbers.")

    node_path = shutil.which("node")
    if node_path is None:
        raise RuntimeError("Node.js is not installed or not on PATH.")

    result = subprocess.run([node_path, "bin/app.js", "register", mobile], cwd=repo, check=False)
    return result.returncode


def save_truecaller_installation_id(installation_id: str, repo_path: str | Path) -> Path:
    """Write a manual installationId config for the vendored truecaller-cli."""
    repo = Path(repo_path)
    normalized = installation_id.strip()
    if len(normalized) < 10:
        raise ValueError("The installationId looks too short.")
    config_path = repo / "config.json"
    config_path.write_text(json.dumps({"installationId": normalized}, indent=2, ensure_ascii=True), encoding="utf-8")
    return config_path


def search_truecaller(phone: str, repo_path: str | Path, timeout_seconds: int = 60) -> TruecallerSearchResult | None:
    """Run the vendored truecaller-cli search command and parse its output."""
    repo = Path(repo_path)
    mobile = _format_truecaller_number(phone)
    if mobile is None:
        return None

    node_path = shutil.which("node")
    if node_path is None:
        raise RuntimeError("Node.js is not installed or not on PATH.")

    completed = subprocess.run(
        [node_path, "bin/app.js", "search", mobile],
        cwd=repo,
        capture_output=True,
        text=True,
        timeout=timeout_seconds,
        check=False,
    )
    output = _strip_ansi("\n".join(part for part in [completed.stdout, completed.stderr] if part).strip())
    if "Installtion Id not found" in output:
        raise RuntimeError("truecaller-cli is not registered yet. Run setup-truecaller first.")
    if "Cannot find module" in output:
        raise RuntimeError("truecaller-cli dependencies are missing. Run npm install in _vendor/truecaller-cli.")
    if "Not Found." in output or "Not a valid Mobile Number" in output or not output:
        return None

    name_match = re.search(r"Name:\s*(.+)", output)
    email_match = re.search(r"Email:\s*(.+)", output)
    if not name_match:
        return None

    return TruecallerSearchResult(
        name=name_match.group(1).strip(),
        email=email_match.group(1).strip() if email_match else None,
        raw_output=output,
    )


def extract_names_from_snippets(
    snippets: list[str],
    *,
    phone: str,
    llm_callable: Callable[[str, list[str]], list[str]] | None = None,
) -> list[str]:
    """Extract candidate names from search snippets."""
    heuristic_candidates: list[str] = []
    for snippet in snippets:
        for match in NAME_PATTERN.findall(snippet):
            candidate = " ".join(match.split())
            if _is_plausible_name(candidate):
                heuristic_candidates.append(candidate)

    candidates = dedupe_strings(heuristic_candidates)

    if llm_callable is not None and snippets:
        try:
            llm_candidates = llm_callable(phone, snippets)
        except Exception:
            llm_candidates = []
        candidates = dedupe_strings([*llm_candidates, *candidates])

    return candidates[:10]


def ask_user(question: str) -> str:
    """Ask the user a question in the console and return the response."""
    return input(f"{question}\n> ").strip()


def save_resolved(contact: ContactRecord, memory_store: MemoryStore) -> ContactRecord:
    """Persist a resolved contact in memory."""
    return memory_store.save_contact(contact)


def finalize(
    phone: str,
    name: str,
    confidence: float,
    source: str,
    reasoning: str,
    memory_store: MemoryStore,
) -> ContactRecord:
    """Create and store the final contact record."""
    record = ContactRecord(
        phone=phone,
        name=name.strip() or "UNKNOWN",
        confidence=clamp_confidence(confidence),
        source=source,
        reasoning=reasoning.strip(),
        last_updated=utc_now_iso(),
    )
    return save_resolved(record, memory_store)


def export_resolved(records: list[ContactRecord], output_path: str | Path) -> Path:
    """Export the final contact set to CSV."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["phone", "name", "confidence", "source", "reasoning", "last_updated"],
        )
        writer.writeheader()
        for record in records:
            writer.writerow(record.model_dump())
    return path


def _load_contacts_from_csv(path: Path, default_region: str) -> list[LoadedContact]:
    contacts: list[LoadedContact] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames:
            phone_columns = [field for field in reader.fieldnames if _looks_like_phone_column(field)]
            for row_number, row in enumerate(reader, start=2):
                values = [row.get(column, "") for column in (phone_columns or [reader.fieldnames[0]])]
                for value in values:
                    normalized = normalize_phone(str(value), default_region)
                    if normalized:
                        contacts.append(
                            LoadedContact(
                                phone=normalized,
                                original_value=str(value).strip(),
                                source_file=str(path),
                                row_number=row_number,
                            )
                        )
        else:
            handle.seek(0)
            fallback_reader = csv.reader(handle)
            for row_number, row in enumerate(fallback_reader, start=1):
                for value in row:
                    normalized = normalize_phone(str(value), default_region)
                    if normalized:
                        contacts.append(
                            LoadedContact(
                                phone=normalized,
                                original_value=str(value).strip(),
                                source_file=str(path),
                                row_number=row_number,
                            )
                        )
    return contacts


def _load_contacts_from_txt(path: Path, default_region: str) -> list[LoadedContact]:
    contacts: list[LoadedContact] = []
    lines = path.read_text(encoding="utf-8").splitlines()
    for row_number, line in enumerate(lines, start=1):
        for candidate in extract_phone_candidates(line):
            normalized = normalize_phone(candidate, default_region)
            if normalized:
                contacts.append(
                    LoadedContact(
                        phone=normalized,
                        original_value=candidate,
                        source_file=str(path),
                        row_number=row_number,
                    )
                )
    return contacts


def _load_contacts_from_xlsx(path: Path, default_region: str) -> list[LoadedContact]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed. Install requirements.txt first.")

    contacts: list[LoadedContact] = []
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers, data_rows, start_row = _split_sheet_rows(rows)
            phone_indices = [index for index, header in enumerate(headers) if _looks_like_phone_column(header)] or [0]
            for offset, row in enumerate(data_rows, start=start_row):
                values = list(row)
                for index in phone_indices:
                    if index >= len(values):
                        continue
                    cell_value = values[index]
                    normalized = normalize_phone(str(cell_value or ""), default_region)
                    if normalized:
                        contacts.append(
                            LoadedContact(
                                phone=normalized,
                                original_value=str(cell_value).strip(),
                                source_file=f"{path}::{worksheet.title}",
                                row_number=offset,
                            )
                        )
    finally:
        workbook.close()
    return contacts


def _load_contacts_from_json(path: Path, default_region: str) -> list[LoadedContact]:
    contacts: list[LoadedContact] = []
    payload = json.loads(path.read_text(encoding="utf-8"))
    raw_values = payload.get("phones", []) if isinstance(payload, dict) else payload
    if not isinstance(raw_values, list):
        return contacts

    for index, item in enumerate(raw_values, start=1):
        value = item.get("phone") if isinstance(item, dict) else item
        normalized = normalize_phone(str(value), default_region)
        if normalized:
            contacts.append(
                LoadedContact(
                    phone=normalized,
                    original_value=str(value).strip(),
                    source_file=str(path),
                    row_number=index,
                )
            )
    return contacts


def _load_contacts_from_vcf(path: Path, default_region: str) -> list[LoadedContact]:
    contacts: list[LoadedContact] = []
    blocks = _split_vcards(path.read_text(encoding="utf-8"))
    for index, block in enumerate(blocks, start=1):
        for tel in _extract_vcard_phones(block):
            normalized = normalize_phone(tel, default_region)
            if normalized:
                contacts.append(
                    LoadedContact(
                        phone=normalized,
                        original_value=tel,
                        source_file=str(path),
                        row_number=index,
                    )
                )
    return contacts


def _load_google_contacts_from_csv(path: Path, default_region: str) -> dict[str, str]:
    matches: dict[str, str] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            return matches

        phone_columns = [
            column
            for column in reader.fieldnames
            if "phone" in column.lower() and ("value" in column.lower() or column.lower() in PHONE_COLUMN_HINTS)
        ]

        for row in reader:
            name = _derive_name_from_contact_row(row)
            if not name:
                continue
            for column in phone_columns:
                normalized = normalize_phone(str(row.get(column, "")), default_region)
                if normalized:
                    matches[normalized] = name
    return matches


def _load_google_contacts_from_xlsx(path: Path, default_region: str) -> dict[str, str]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed. Install requirements.txt first.")

    matches: dict[str, str] = {}
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers, data_rows, _ = _split_sheet_rows(rows)
            for row in data_rows:
                row_map = _row_to_dict(headers, row)
                name = _derive_name_from_contact_row(row_map)
                if not name:
                    name = _derive_name_from_generic_row(headers, row)
                if not name:
                    continue
                for header, value in row_map.items():
                    if not _looks_like_phone_column(header):
                        continue
                    normalized = normalize_phone(str(value or ""), default_region)
                    if normalized:
                        matches[normalized] = name
    finally:
        workbook.close()
    return matches


def _load_google_contacts_from_vcf(path: Path, default_region: str) -> dict[str, str]:
    matches: dict[str, str] = {}
    blocks = _split_vcards(path.read_text(encoding="utf-8"))
    for block in blocks:
        name = _extract_vcard_name(block)
        if not name:
            continue
        for tel in _extract_vcard_phones(block):
            normalized = normalize_phone(tel, default_region)
            if normalized:
                matches[normalized] = name
    return matches


def _split_vcards(text: str) -> list[list[str]]:
    blocks: list[list[str]] = []
    current: list[str] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if line.upper() == "BEGIN:VCARD":
            current = [line]
        elif line.upper() == "END:VCARD":
            current.append(line)
            blocks.append(current)
            current = []
        elif current:
            current.append(line)
    return blocks


def _extract_vcard_name(block: list[str]) -> str:
    for line in block:
        if line.upper().startswith("FN"):
            return line.split(":", 1)[-1].strip()
    for line in block:
        if line.upper().startswith("N:"):
            parts = [part.strip() for part in line.split(":", 1)[-1].split(";")]
            ordered = [parts[1], parts[2], parts[0]] if len(parts) >= 3 else parts
            name = " ".join(part for part in ordered if part)
            if name:
                return name
    return ""


def _extract_vcard_phones(block: list[str]) -> list[str]:
    phones: list[str] = []
    for line in block:
        if line.upper().startswith("TEL"):
            phones.append(line.split(":", 1)[-1].strip())
    return phones


def _derive_name_from_contact_row(row: dict[str, Any]) -> str:
    direct_name = str(row.get("Name") or row.get("Full Name") or row.get("Public Name") or "").strip()
    if direct_name:
        return direct_name

    parts = [
        str(row.get("Given Name") or "").strip(),
        str(row.get("Additional Name") or "").strip(),
        str(row.get("Family Name") or "").strip(),
    ]
    return " ".join(part for part in parts if part).strip()


def _looks_like_phone_column(column_name: str) -> bool:
    normalized = column_name.strip().lower()
    if normalized in PHONE_COLUMN_HINTS:
        return True
    return (
        "phone" in normalized
        or "mobile" in normalized
        or "telephone" in normalized
        or "contact no" in normalized
        or "contact number" in normalized
        or "mobile no" in normalized
        or "mobile number" in normalized
    )


def _looks_like_name_column(column_name: str) -> bool:
    normalized = column_name.strip().lower()
    return normalized in {
        "name",
        "full name",
        "public name",
        "contact name",
        "given name",
        "family name",
        "first name",
        "last name",
    }


def _split_sheet_rows(rows: list[tuple[Any, ...]]) -> tuple[list[str], list[tuple[Any, ...]], int]:
    first_row = rows[0]
    if _row_looks_like_header(first_row):
        headers = [str(cell).strip() if cell is not None else "" for cell in first_row]
        return headers, rows[1:], 2

    width = max(len(first_row), 1)
    headers = [f"column_{index + 1}" for index in range(width)]
    return headers, rows, 1


def _row_looks_like_header(row: tuple[Any, ...]) -> bool:
    values = [str(cell).strip() for cell in row if cell not in (None, "")]
    if not values:
        return False
    return any(_looks_like_phone_column(value) or _looks_like_name_column(value) for value in values)


def _row_to_dict(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    values = list(row)
    row_map: dict[str, Any] = {}
    for index, header in enumerate(headers):
        row_map[header] = values[index] if index < len(values) else None
    return row_map


def _derive_name_from_generic_row(headers: list[str], row: tuple[Any, ...]) -> str:
    values = list(row)
    for index, header in enumerate(headers):
        if not _looks_like_name_column(header):
            continue
        if index >= len(values):
            continue
        candidate = str(values[index] or "").strip()
        if candidate:
            return candidate
    return ""


def _is_plausible_name(candidate: str) -> bool:
    tokens = candidate.split()
    if not tokens or len(tokens) > 3:
        return False
    if any(token in GENERIC_NAME_TERMS for token in tokens):
        return False
    if not any(len(token) > 2 for token in tokens):
        return False
    return True


def _format_truecaller_number(phone: str) -> str | None:
    digits = re.sub(r"\D", "", phone or "")
    if len(digits) == 12 and digits.startswith("91"):
        return digits[-10:]
    if len(digits) == 10:
        return digits
    return None


def _strip_ansi(text: str) -> str:
    return re.sub(r"\x1b\[[0-9;]*m", "", text)
